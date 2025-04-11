import random
import json
import logging
import pymongo
import os
from bson import ObjectId
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()
def initialize_client():
    """Initialize MongoDB Client"""
    mongo_host = os.getenv('MONGO_HOST', 'localhost')
    mongo_port = os.getenv('MONGO_PORT', '27017')
    mongo_user = os.getenv('MONGO_USER')
    mongo_pass = os.getenv('MONGO_PWD')
    mongo_auth_src = os.getenv('MONGO_AUTH_SRC', 'admin')

    logging.info(f"Connecting to MongoDB at {mongo_host}:{mongo_port}")

    try:
        client = pymongo.MongoClient(
            host=[f'{mongo_host}:{mongo_port}'],
            username=mongo_user,
            password=mongo_pass,
            authSource=mongo_auth_src,
            authMechanism='SCRAM-SHA-256',
            maxPoolSize=100,
            serverSelectionTimeoutMS=5000  # Avoid indefinite hanging
        )
        return client
    except Exception as e:
        logging.error(f"MongoDB connection failed: {e}")
        raise


def query_publications(ids):
    client = initialize_client()
    db = client['oeb-research-software']
    publications = []
    for id in ids:
        entry = db.publications_metadata.find_one({"_id": ObjectId(id['$oid'])})
        if entry:
            publications.append(entry['data'])

    return publications


def build_instances_keys_dict(json_entries):
    """Create a mapping of instance IDs to their respective instance data."""
    with open(json_entries, "r", encoding="utf-8") as f:
        data = json.load(f)

    instances_keys = {}
    for key, value in data.items():
        instances = value.get("instances", [])
        for instance in instances:
            instances_keys[instance["_id"]] = instance
    return instances_keys



def random_items(data, n=100):
    random_keys = random.sample(list(data.keys()), n)
    random_items = {key: data[key] for key in random_keys}

    return random_items


def get_repos(entry):
    repos = []
    if entry['data'].get('repository'):
        for repo in entry['data']['repository']:
            if repo.get('url'):
                repos.append(repo['url'])
    
    if len(repos) == 0:
        return ''

    elif len(repos) == 1:
        repo_string = f'=HYPERLINK("{repos[0]}","{repos[0]}")'
        return repo_string
    
    else: 
        repo_str = "\n".join(repos)
        return repo_str

def get_webpage(entry):
    webs = entry['data']['webpage']

    if len(webs) == 0:
        return ''

    elif len(webs) == 1:
        web_string = f'=HYPERLINK("{webs[0]}","{webs[0]}")'
        return web_string
    
    else: 
        web_str = "\n".join(webs)
        return web_str
    
def get_authors(entry):
    authors = entry['data']['authors']
    n = 0

    full_string = ""
    for author in authors:
        item = ''
        if author.get('name', ''):
            item = f"{author.get('name', '')}"
        
        if author.get('email', ''):
            item = f"{item} <{author.get('email', '')}>"
        
        if n == 0:
            full_string = item

        else:
            full_string = f"{full_string}\n{item}"
        n += 1
    
    return full_string

def get_publication(entry):
    pubs = entry['data']['publication']
    publications = query_publications(pubs)

    full_string = ''
    if len(publications) == 0:
        return ''
    
    elif len(publications) == 1:
        publication = publications[0]
        full_string = f"{publication.get('title', '')} ({publication.get('year', '')}),{publication.get('doi','-')}, {publication.get('url','-')}"
    
    else:
        for publication in publications:
            item = f"{publication.get('title', '')} ({publication.get('year', '')}),{publication.get('doi','-')}, {publication.get('url','-')}"
            if full_string == '':
                full_string = item
            else:
                full_string = f"{full_string}\n{item}"
    
    return full_string
    
    
def get_license(entry):
    licenses = entry['data']['license']
    n = 0

    full_string = ""
    for license in licenses:
        item = f"{license.get('name', '')}"
        if n == 0:
            full_string = item
        else:
            full_string = f"{full_string}\n{item}"
        n += 1
    
    return full_string


def build_rows_dict(random_items, instances_dict):
    """Create a list of dictionaries for each entry."""

    pairs_to_evaluate = []
    for key,value in random_items.items():
        pair = {}
        pair['entry_id'] = key
        #print(f"There are {len(value['disconnected'])} disconnected and {len(value['remaining'])} remaining entries.")

        entry_1_id = value['disconnected'][0]['id']
        if value.get('remaining'):
            entry_2_id = value['remaining'][0]['id']
        else:
            entry_2_id = value['disconnected'][1]['id']

        entry_1 = instances_dict[entry_1_id]
        entry_2 = instances_dict[entry_2_id]

        pair['entry_1_id'] = entry_1_id
        pair['entry_1_source'] = entry_1['data']['source'][0]
        pair['entry_1_name'] = entry_1['data']['name']
        pair['entry_1_type'] = entry_1['data']['type']
        pair['entry_1_description'] = entry_1['data']['description'][0] if entry_1['data'].get('description') else ""
        pair['entry_1_repository'] = get_repos(entry_1)
        pair['entry_1_webpage'] = get_webpage(entry_1)
        pair['entry_1_authors'] = get_authors(entry_1)
        pair['entry_1_publication'] = get_publication(entry_1)
        pair['entry_1_license'] = get_license(entry_1)

        pair['entry_2_id'] = entry_2_id
        pair['entry_2_source'] = entry_2['data']['source'][0]
        pair['entry_2_name'] = entry_2['data']['name']
        pair['entry_2_type'] = entry_2['data']['type']
        pair['entry_2_description'] = entry_2['data']['description'][0] if entry_2['data'].get('description') else ""
        pair['entry_2_repository'] = get_repos(entry_2)
        pair['entry_2_webpage'] = get_webpage(entry_2)
        pair['entry_2_authors'] = get_authors(entry_2)
        pair['entry_2_publication'] = get_publication(entry_2)
        pair['entry_2_license'] = get_license(entry_2)

        pairs_to_evaluate.append(pair)

    return pairs_to_evaluate


if __name__ == "__main__":
    # Load from file
    json_path = "scripts/data/disconnected_entries.json"
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    instances_dict = build_instances_keys_dict("scripts/data/grouped_entries.json")

    random_items = random_items(data, n=100)
    pairs_to_evaluate = build_rows_dict(random_items, instances_dict)

    df = pd.DataFrame(pairs_to_evaluate)
    df["human_verdict"] = ""
    df["human_confidence"] = ""
    df["human_rationale"] = ""
    df["llm_verdict"] = ""
    df["llm_confidence"] = ""
    df["llm_key_features"] = ""
    df["llm_rationale"] = ""

    output_path = "software_disambiguation_filled_template.xlsx"
    df.to_excel(output_path, index=False)
    print(f"Saved: {output_path}")

    # ---- formatting the whole spreadsheet ----
    # Load workbook
    wb = load_workbook(output_path)
    ws = wb.active

    # Loop through columns to adjust:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                # Measure length of cell for width estimate
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        # Set width (with padding factor)
        adjusted_width = min(max_length * 1.1, 80)  # Cap to avoid excessive width
        ws.column_dimensions[col_letter].width = adjusted_width

    # Get column letters for target columns
    header = [cell.value for cell in ws[1]]
    veredict_col = header.index("human_verdict") + 1
    confidence_col = header.index("human_confidence") + 1

    veredict_letter = ws.cell(row=1, column=veredict_col).column_letter
    confidence_letter = ws.cell(row=1, column=confidence_col).column_letter

    # Create data validation dropdowns
    veredict_dropdown = DataValidation(
        type="list",
        formula1='"same,different,unclear"',
        showDropDown=True
    )
    confidence_dropdown = DataValidation(
        type="list",
        formula1='"high,medium,low"',
        showDropDown=True
    )

    # Add dropdowns to the sheet
    ws.add_data_validation(veredict_dropdown)
    ws.add_data_validation(confidence_dropdown)

    # Apply dropdowns to target columns (from row 2 down)
    veredict_range = f"{veredict_letter}2:{veredict_letter}{ws.max_row}"
    confidence_range = f"{confidence_letter}2:{confidence_letter}{ws.max_row}"

    veredict_dropdown.add(veredict_range)
    confidence_dropdown.add(confidence_range)

    # Save changes
    wb.save(output_path)





    

    